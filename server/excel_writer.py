from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel(data, metadata, output_stream):
    """
    Generates an Excel file from transaction data.
    data: List of dictionaries or lists representing rows.
    metadata: Dictionary containing headers and other info.
    output_stream: BytesIO stream to write to.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mutasi Rekening"

    headers = metadata.get("headers", [])
    
    # Header styling
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald Green
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Formatting numbers if it's a float or int
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            
            # General alignment
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="top")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50) # Cap at 50

    wb.save(output_stream)
